

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import Workbook

class TestDataLoader:
    def __init__(self, excel_file):
        self.excel_file = excel_file

    def read_test_data(self, sheet_name):
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook[sheet_name]

            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            workbook.close()
            return data

        except FileNotFoundError:
            print(f"Error: File '{self.excel_file}' not found.")
            return []

        except InvalidFileException:
            print(f"Error: Invalid Excel file format for '{self.excel_file}'.")
            return []

        except Exception as e:
            print(f"Error: An unexpected error occurred - {e}")
            return []

    def modify_and_save(self, sheet_name):
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook[sheet_name]

            # Make modifications to the specified sheet
            # For example, appending data to an existing sheet
            sheet.append(["Modified", "Data"])

            # Save the modified workbook with changes
            modified_excel_file = "ModifiedData.xlsx"
            workbook.save(filename=modified_excel_file)

            workbook.close()  # Close the workbook after saving
            return modified_excel_file  # Return the file path of the modified workbook

        except FileNotFoundError:
            print(f"Error: File '{self.excel_file}' not found.")
            return ""

        except InvalidFileException:
            print(f"Error: Invalid Excel file format for '{self.excel_file}'.")
            return ""

        except Exception as e:
            print(f"Error: An unexpected error occurred - {e}")
            return ""

# Instantiate TestDataLoader with the path to your Excel file
test_data_loader = TestDataLoader("C:/Users/rajah/PycharmProjects/ProjectZoobook/testdata/test.xlsx")

# Read test data from a specific sheet
test_data = test_data_loader.read_test_data("test")

if test_data:
    print("Test data successfully loaded.")
    # Call modify_and_save method to modify the Excel file
    modified_file_path = test_data_loader.modify_and_save("test")

    if modified_file_path:
        print(f"Modified workbook saved as '{modified_file_path}'")
    else:
        print("Failed to process and save the workbook")
else:
    print("Failed to load test data.")
