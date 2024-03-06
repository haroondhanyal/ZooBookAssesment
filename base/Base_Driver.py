from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook  # Library for working with Excel

class BaseDriver:
    def __init__(self):
        self.driver = webdriver.Chrome()
        self.wait = WebDriverWait(self.driver, 10)

    def navigate_to_url(self, url):
        self.driver.get(url)

    def find_element(self, by, value):
        return self.wait.until(EC.presence_of_element_located((by, value)))

    def find_elements(self, by, value):
        return self.wait.until(EC.presence_of_all_elements_located((by, value)))

    def click_element(self, by, value):
        element = self.find_element(by, value)
        element.click()

    def input_text(self, by, value, text):
        element = self.find_element(by, value)
        element.clear()
        element.send_keys(text)

    def quit_driver(self):
        self.driver.quit()

    def read_excel_data(self, file_path, sheet_name):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data

# Example Usage:
# Instantiate the BaseDriver
base_driver = BaseDriver()

# Navigate to a URL
base_driver.navigate_to_url("https://staging.zoobooksystems.com/")

# Read test data from Excel
test_data = base_driver.read_excel_data("testdata.xlsx", "login_data")

# Iterate through test data and perform actions
for data_set in test_data:
    base_driver.input_text(By.XPATH, "//input[@name='zb-username']", data_set[0])
    base_driver.input_text(By.XPATH, "//input[@name='zb-password']", data_set[1])
    base_driver.click_element(By.XPATH, "//button[@class='btn btn-lg btn-success btn-block login-button']")

    # Additional actions or validations based on the test case

    # Clear inputs for the next iteration
    base_driver.input_text(By.XPATH, "//input[@name='zb-username']", "")
    base_driver.input_text(By.XPATH, "//input[@name='zb-password']", "")

# Quit the WebDriver when done
base_driver.quit_driver()
